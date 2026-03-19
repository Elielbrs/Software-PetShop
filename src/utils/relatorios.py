"""
Classe de Relatórios
"""
from datetime import datetime
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..models.database import Database


class Relatorios:
    """Classe para gerar relatórios diversos"""
    
    @staticmethod
    def faturamento_diario():
        """Calcula o faturamento do dia"""
        try:
            conn = Database.get_connection()
            cursor = conn.cursor()
            data_atual = datetime.now().strftime("%Y-%m-%d")
            
            cursor.execute("""
                SELECT SUM(s.preco) 
                FROM agendamentos a
                JOIN servicos s ON a.id_servico = s.id
                WHERE a.data_hora LIKE ? AND a.status = 'concluído'
            """, (f"{data_atual}%",))
            
            serviços = cursor.fetchone()[0] or 0.0
            
            cursor.execute("""
                SELECT SUM(v.quantidade * v.preco_unitario)
                FROM vendas v
                WHERE v.data_venda LIKE ?
            """, (f"{data_atual}%",))
            
            vendas = cursor.fetchone()[0] or 0.0
            total = serviços + vendas
            
            print("\n" + "="*60)
            print(f"FATURAMENTO DO DIA {data_atual}".center(60))
            print("="*60)
            print(f"Serviços: R${serviços:.2f}")
            print(f"Vendas:   R${vendas:.2f}")
            print("-"*60)
            print(f"TOTAL:    R${total:.2f}")
            print("="*60 + "\n")
            
            conn.close()
            return total
        except Exception as e:
            print(f"✗ Erro ao calcular faturamento diário: {e}")
            return 0.0
    
    @staticmethod
    def faturamento_mensal():
        """Calcula o faturamento do mês"""
        try:
            conn = Database.get_connection()
            cursor = conn.cursor()
            mes_atual = datetime.now().strftime("%Y-%m")
            
            cursor.execute("""
                SELECT SUM(s.preco) 
                FROM agendamentos a
                JOIN servicos s ON a.id_servico = s.id
                WHERE a.data_hora LIKE ? AND a.status = 'concluído'
            """, (f"{mes_atual}%",))
            
            serviços = cursor.fetchone()[0] or 0.0
            
            cursor.execute("""
                SELECT SUM(v.quantidade * v.preco_unitario)
                FROM vendas v
                WHERE v.data_venda LIKE ?
            """, (f"{mes_atual}%",))
            
            vendas = cursor.fetchone()[0] or 0.0
            total = serviços + vendas
            
            print("\n" + "="*60)
            print(f"FATURAMENTO DO MÊS {mes_atual}".center(60))
            print("="*60)
            print(f"Serviços: R${serviços:.2f}")
            print(f"Vendas:   R${vendas:.2f}")
            print("-"*60)
            print(f"TOTAL:    R${total:.2f}")
            print("="*60 + "\n")
            
            conn.close()
            return total
        except Exception as e:
            print(f"✗ Erro ao calcular faturamento mensal: {e}")
            return 0.0
    
    @staticmethod
    def faturamento_por_servico():
        """Faturamento por serviço do mês"""
        try:
            conn = Database.get_connection()
            cursor = conn.cursor()
            mes_atual = datetime.now().strftime("%Y-%m")
            
            cursor.execute("""
                SELECT s.nome_servico, COUNT(a.id) as total_servicos, SUM(s.preco) as faturamento
                FROM agendamentos a
                JOIN servicos s ON a.id_servico = s.id
                WHERE a.data_hora LIKE ? AND a.status = 'concluído'
                GROUP BY s.nome_servico
                ORDER BY faturamento DESC
            """, (f"{mes_atual}%",))
            
            resultados = cursor.fetchall()
            
            print("\n" + "="*70)
            print(f"FATURAMENTO POR SERVIÇO - {mes_atual}".center(70))
            print("="*70)
            for resultado in resultados:
                print(f"Serviço: {resultado[0]:<30} | Qtd: {resultado[1]:<4} | Faturamento: R${resultado[2]:.2f}")
            print("="*70 + "\n")
            
            conn.close()
            return resultados
        except Exception as e:
            print(f"✗ Erro ao calcular faturamento por serviço: {e}")
            return []
    
    @staticmethod
    def faturamento_por_produto():
        """Faturamento por produto do mês"""
        try:
            conn = Database.get_connection()
            cursor = conn.cursor()
            mes_atual = datetime.now().strftime("%Y-%m")
            
            cursor.execute("""
                SELECT p.nome, SUM(v.quantidade) as total_vendido, SUM(v.quantidade * v.preco_unitario) as faturamento
                FROM vendas v
                JOIN produtos p ON v.id_produto = p.id
                WHERE v.data_venda LIKE ?
                GROUP BY p.nome
                ORDER BY faturamento DESC
            """, (f"{mes_atual}%",))
            
            resultados = cursor.fetchall()
            
            print("\n" + "="*70)
            print(f"FATURAMENTO POR PRODUTO - {mes_atual}".center(70))
            print("="*70)
            for resultado in resultados:
                print(f"Produto: {resultado[0]:<35} | Qtd: {resultado[1]:<5} | Faturamento: R${resultado[2]:.2f}")
            print("="*70 + "\n")
            
            conn.close()
            return resultados
        except Exception as e:
            print(f"✗ Erro ao calcular faturamento por produto: {e}")
            return []
    
    @staticmethod
    def produto_mais_vendido():
        """Produto mais vendido do mês"""
        try:
            conn = Database.get_connection()
            cursor = conn.cursor()
            mes_atual = datetime.now().strftime("%Y-%m")
            
            cursor.execute("""
                SELECT p.nome, SUM(v.quantidade) as total
                FROM vendas v
                JOIN produtos p ON v.id_produto = p.id
                WHERE v.data_venda LIKE ?
                GROUP BY p.nome
                ORDER BY total DESC
                LIMIT 1
            """, (f"{mes_atual}%",))
            
            resultado = cursor.fetchone()
            
            if resultado:
                print(f"\n✓ Produto mais vendido em {mes_atual}: {resultado[0]} ({resultado[1]} unidades)\n")
            else:
                print(f"\n⚠ Nenhuma venda registrada em {mes_atual}\n")
            
            conn.close()
            return resultado
        except Exception as e:
            print(f"✗ Erro ao buscar produto mais vendido: {e}")
            return None
    
    @staticmethod
    def gerar_relatorio_excel(arquivo_saida="relatorio.xlsx"):
        """Gera relatório completo em arquivo Excel"""
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove a planilha padrão
            
            # Estilos
            cabecalho_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cabecalho_font = Font(bold=True, color="FFFFFF", size=12)
            total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            total_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            conn = Database.get_connection()
            cursor = conn.cursor()
            
            # Data atual
            data_relatorio = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            mes_ano = datetime.now().strftime("%B/%Y")
            
            # ========== ABA 1: FATURAMENTO MENSAL ==========
            ws_faturamento = wb.create_sheet("Faturamento Mensal")
            
            # Cabeçalho
            ws_faturamento.merge_cells('A1:D1')
            ws_faturamento['A1'] = f"RELATÓRIO DE FATURAMENTO - {mes_ano}"
            ws_faturamento['A1'].font = Font(bold=True, size=14)
            ws_faturamento['A1'].alignment = Alignment(horizontal="center", vertical="center")
            
            # Data do relatório
            ws_faturamento.merge_cells('A2:D2')
            ws_faturamento['A2'] = f"Gerado em: {data_relatorio}"
            ws_faturamento['A2'].font = Font(italic=True, size=10)
            ws_faturamento['A2'].alignment = Alignment(horizontal="center")
            
            ws_faturamento.append([])
            
            # Cabeçalhos das colunas
            headers = ["Categoria", "Valor (R$)", "Percentual", "Origem"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_faturamento.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = cabecalho_fill
                cell.font = cabecalho_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            # Dados de faturamento
            mes_atual = datetime.now().strftime("%Y-%m")
            
            cursor.execute("""
                SELECT SUM(s.preco) 
                FROM agendamentos a
                JOIN servicos s ON a.id_servico = s.id
                WHERE a.data_hora LIKE ? AND a.status = 'concluído'
            """, (f"{mes_atual}%",))
            faturamento_servicos = cursor.fetchone()[0] or 0.0
            
            cursor.execute("""
                SELECT SUM(v.quantidade * v.preco_unitario)
                FROM vendas v
                WHERE v.data_venda LIKE ?
            """, (f"{mes_atual}%",))
            faturamento_vendas = cursor.fetchone()[0] or 0.0
            
            total_faturamento = faturamento_servicos + faturamento_vendas
            
            # Adiciona linhas de dados
            row_num = 5
            dados = [
                ("Serviços", faturamento_servicos),
                ("Vendas de Produtos", faturamento_vendas)
            ]
            
            for categoria, valor in dados:
                ws_faturamento.cell(row=row_num, column=1).value = categoria
                ws_faturamento.cell(row=row_num, column=2).value = valor
                percentual = (valor / total_faturamento * 100) if total_faturamento > 0 else 0
                ws_faturamento.cell(row=row_num, column=3).value = f"{percentual:.1f}%"
                ws_faturamento.cell(row=row_num, column=4).value = "Interno"
                
                for col in range(1, 5):
                    ws_faturamento.cell(row=row_num, column=col).border = border
                    if col == 2:
                        ws_faturamento.cell(row=row_num, column=col).number_format = 'R$ #,##0.00'
                
                row_num += 1
            
            # Linha de total
            ws_faturamento.cell(row=row_num, column=1).value = "TOTAL"
            ws_faturamento.cell(row=row_num, column=2).value = total_faturamento
            ws_faturamento.cell(row=row_num, column=3).value = "100.0%"
            ws_faturamento.cell(row=row_num, column=4).value = ""
            
            for col in range(1, 5):
                cell = ws_faturamento.cell(row=row_num, column=col)
                cell.fill = total_fill
                cell.font = total_font
                cell.border = border
                if col == 2:
                    cell.number_format = 'R$ #,##0.00'
            
            # Ajusta largura das colunas
            ws_faturamento.column_dimensions['A'].width = 25
            ws_faturamento.column_dimensions['B'].width = 15
            ws_faturamento.column_dimensions['C'].width = 12
            ws_faturamento.column_dimensions['D'].width = 15
            
            # ========== ABA 2: FATURAMENTO POR SERVIÇO ==========
            ws_servicos = wb.create_sheet("Faturamento por Serviço")
            
            ws_servicos.merge_cells('A1:E1')
            ws_servicos['A1'] = f"FATURAMENTO POR SERVIÇO - {mes_ano}"
            ws_servicos['A1'].font = Font(bold=True, size=14)
            ws_servicos['A1'].alignment = Alignment(horizontal="center")
            
            ws_servicos.merge_cells('A2:E2')
            ws_servicos['A2'] = f"Gerado em: {data_relatorio}"
            ws_servicos['A2'].font = Font(italic=True, size=10)
            ws_servicos['A2'].alignment = Alignment(horizontal="center")
            
            ws_servicos.append([])
            
            headers = ["Serviço", "Quantidade", "Faturamento (R$)", "Percentual", "Preço Médio"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_servicos.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = cabecalho_fill
                cell.font = cabecalho_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            cursor.execute("""
                SELECT s.nome_servico, COUNT(a.id) as total_servicos, SUM(s.preco) as faturamento
                FROM agendamentos a
                JOIN servicos s ON a.id_servico = s.id
                WHERE a.data_hora LIKE ? AND a.status = 'concluído'
                GROUP BY s.nome_servico
                ORDER BY faturamento DESC
            """, (f"{mes_atual}%",))
            
            resultados_servicos = cursor.fetchall()
            total_servicos_valor = sum([r[2] for r in resultados_servicos]) if resultados_servicos else 0
            
            row_num = 5
            for resultado in resultados_servicos:
                ws_servicos.cell(row=row_num, column=1).value = resultado[0]
                ws_servicos.cell(row=row_num, column=2).value = resultado[1]
                ws_servicos.cell(row=row_num, column=3).value = resultado[2]
                percentual = (resultado[2] / total_servicos_valor * 100) if total_servicos_valor > 0 else 0
                ws_servicos.cell(row=row_num, column=4).value = f"{percentual:.1f}%"
                preco_medio = resultado[2] / resultado[1] if resultado[1] > 0 else 0
                ws_servicos.cell(row=row_num, column=5).value = preco_medio
                
                for col in range(1, 6):
                    ws_servicos.cell(row=row_num, column=col).border = border
                    if col in [3, 5]:
                        ws_servicos.cell(row=row_num, column=col).number_format = 'R$ #,##0.00'
                
                row_num += 1
            
            # Total
            ws_servicos.cell(row=row_num, column=1).value = "TOTAL"
            ws_servicos.cell(row=row_num, column=2).value = sum([r[1] for r in resultados_servicos]) if resultados_servicos else 0
            ws_servicos.cell(row=row_num, column=3).value = total_servicos_valor
            ws_servicos.cell(row=row_num, column=4).value = "100.0%"
            ws_servicos.cell(row=row_num, column=5).value = ""
            
            for col in range(1, 6):
                cell = ws_servicos.cell(row=row_num, column=col)
                cell.fill = total_fill
                cell.font = total_font
                cell.border = border
                if col in [3, 5]:
                    cell.number_format = 'R$ #,##0.00'
            
            ws_servicos.column_dimensions['A'].width = 30
            ws_servicos.column_dimensions['B'].width = 12
            ws_servicos.column_dimensions['C'].width = 18
            ws_servicos.column_dimensions['D'].width = 12
            ws_servicos.column_dimensions['E'].width = 15
            
            # ========== ABA 3: FATURAMENTO POR PRODUTO ==========
            ws_produtos = wb.create_sheet("Faturamento por Produto")
            
            ws_produtos.merge_cells('A1:E1')
            ws_produtos['A1'] = f"FATURAMENTO POR PRODUTO - {mes_ano}"
            ws_produtos['A1'].font = Font(bold=True, size=14)
            ws_produtos['A1'].alignment = Alignment(horizontal="center")
            
            ws_produtos.merge_cells('A2:E2')
            ws_produtos['A2'] = f"Gerado em: {data_relatorio}"
            ws_produtos['A2'].font = Font(italic=True, size=10)
            ws_produtos['A2'].alignment = Alignment(horizontal="center")
            
            ws_produtos.append([])
            
            headers = ["Produto", "Quantidade", "Faturamento (R$)", "Percentual", "Preço Médio"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_produtos.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = cabecalho_fill
                cell.font = cabecalho_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            cursor.execute("""
                SELECT p.nome, SUM(v.quantidade) as total_vendido, SUM(v.quantidade * v.preco_unitario) as faturamento
                FROM vendas v
                JOIN produtos p ON v.id_produto = p.id
                WHERE v.data_venda LIKE ?
                GROUP BY p.nome
                ORDER BY faturamento DESC
            """, (f"{mes_atual}%",))
            
            resultados_produtos = cursor.fetchall()
            total_produtos_valor = sum([r[2] for r in resultados_produtos]) if resultados_produtos else 0
            
            row_num = 5
            for resultado in resultados_produtos:
                ws_produtos.cell(row=row_num, column=1).value = resultado[0]
                ws_produtos.cell(row=row_num, column=2).value = resultado[1]
                ws_produtos.cell(row=row_num, column=3).value = resultado[2]
                percentual = (resultado[2] / total_produtos_valor * 100) if total_produtos_valor > 0 else 0
                ws_produtos.cell(row=row_num, column=4).value = f"{percentual:.1f}%"
                preco_medio = resultado[2] / resultado[1] if resultado[1] > 0 else 0
                ws_produtos.cell(row=row_num, column=5).value = preco_medio
                
                for col in range(1, 6):
                    ws_produtos.cell(row=row_num, column=col).border = border
                    if col in [3, 5]:
                        ws_produtos.cell(row=row_num, column=col).number_format = 'R$ #,##0.00'
                
                row_num += 1
            
            # Total
            ws_produtos.cell(row=row_num, column=1).value = "TOTAL"
            ws_produtos.cell(row=row_num, column=2).value = sum([r[1] for r in resultados_produtos]) if resultados_produtos else 0
            ws_produtos.cell(row=row_num, column=3).value = total_produtos_valor
            ws_produtos.cell(row=row_num, column=4).value = "100.0%"
            ws_produtos.cell(row=row_num, column=5).value = ""
            
            for col in range(1, 6):
                cell = ws_produtos.cell(row=row_num, column=col)
                cell.fill = total_fill
                cell.font = total_font
                cell.border = border
                if col in [3, 5]:
                    cell.number_format = 'R$ #,##0.00'
            
            ws_produtos.column_dimensions['A'].width = 35
            ws_produtos.column_dimensions['B'].width = 12
            ws_produtos.column_dimensions['C'].width = 18
            ws_produtos.column_dimensions['D'].width = 12
            ws_produtos.column_dimensions['E'].width = 15
            
            # ========== ABA 4: AGENDAMENTOS DO MÊS ==========
            ws_agendamentos = wb.create_sheet("Agendamentos")
            
            ws_agendamentos.merge_cells('A1:F1')
            ws_agendamentos['A1'] = f"AGENDAMENTOS DO MÊS - {mes_ano}"
            ws_agendamentos['A1'].font = Font(bold=True, size=14)
            ws_agendamentos['A1'].alignment = Alignment(horizontal="center")
            
            ws_agendamentos.merge_cells('A2:F2')
            ws_agendamentos['A2'] = f"Gerado em: {data_relatorio}"
            ws_agendamentos['A2'].font = Font(italic=True, size=10)
            ws_agendamentos['A2'].alignment = Alignment(horizontal="center")
            
            ws_agendamentos.append([])
            
            headers = ["Data", "Horário", "Pet", "Dono", "Serviço", "Status"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_agendamentos.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = cabecalho_fill
                cell.font = cabecalho_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            cursor.execute("""
                SELECT 
                    DATE(a.data_hora) as data,
                    TIME(a.data_hora) as hora,
                    p.nome as pet,
                    u.nome as dono,
                    s.nome_servico,
                    a.status
                FROM agendamentos a
                JOIN pets p ON a.id_pet = p.id
                JOIN usuarios u ON p.dono_id = u.id
                JOIN servicos s ON a.id_servico = s.id
                WHERE a.data_hora LIKE ?
                ORDER BY a.data_hora DESC
            """, (f"{mes_atual}%",))
            
            agendamentos = cursor.fetchall()
            
            row_num = 5
            total_agendamentos = 0
            for agendamento in agendamentos:
                ws_agendamentos.cell(row=row_num, column=1).value = agendamento[0]
                ws_agendamentos.cell(row=row_num, column=2).value = agendamento[1]
                ws_agendamentos.cell(row=row_num, column=3).value = agendamento[2]
                ws_agendamentos.cell(row=row_num, column=4).value = agendamento[3]
                ws_agendamentos.cell(row=row_num, column=5).value = agendamento[4]
                
                status = agendamento[5]
                ws_agendamentos.cell(row=row_num, column=6).value = status
                
                # Colorir status
                status_cell = ws_agendamentos.cell(row=row_num, column=6)
                if status == "concluído":
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006100", bold=True)
                elif status == "cancelado":
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006", bold=True)
                elif status == "pendente":
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    status_cell.font = Font(color="9C6500")
                
                for col in range(1, 7):
                    ws_agendamentos.cell(row=row_num, column=col).border = border
                
                row_num += 1
                total_agendamentos += 1
            
            # Total de agendamentos
            ws_agendamentos.cell(row=row_num, column=1).value = f"Total: {total_agendamentos} agendamentos"
            ws_agendamentos.cell(row=row_num, column=1).font = total_font
            ws_agendamentos.merge_cells(f'A{row_num}:F{row_num}')
            ws_agendamentos.cell(row=row_num, column=1).fill = total_fill
            ws_agendamentos.cell(row=row_num, column=1).border = border
            
            ws_agendamentos.column_dimensions['A'].width = 12
            ws_agendamentos.column_dimensions['B'].width = 10
            ws_agendamentos.column_dimensions['C'].width = 20
            ws_agendamentos.column_dimensions['D'].width = 20
            ws_agendamentos.column_dimensions['E'].width = 20
            ws_agendamentos.column_dimensions['F'].width = 12
            
            # Salva o arquivo
            diretorio = os.path.dirname(arquivo_saida) or "."
            if not os.path.exists(diretorio):
                os.makedirs(diretorio)
            
            wb.save(arquivo_saida)
            conn.close()
            
            print(f"\n✓ Relatório Excel gerado com sucesso: {os.path.abspath(arquivo_saida)}\n")
            return True
            
        except Exception as e:
            print(f"✗ Erro ao gerar relatório Excel: {e}")
            return False
